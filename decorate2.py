from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

wb = load_workbook('test5.xlsx')
ws = wb.active
index = []
tem = []
for i,cell in enumerate(ws['A']):
    print(i,cell.value)
    flag = cell.value not in tem
    if flag:
        tem.append(cell.value)
    else:
        index.append(i)

fill = PatternFill("solid", fgColor="FF0000")

for i,cell in enumerate(ws['A']):
    if i in index:
        cell.fill = fill
wb.save('test5.xlsx')