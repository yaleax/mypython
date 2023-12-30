def open():
    from openpyxl import load_workbook
    wb = load_workbook('test1.xlsx')
    sh1 = wb['Sheet1']
    shi2 = wb.active
    print(sh1 == shi2)

def show():
    from openpyxl import load_workbook
    wb = load_workbook('test1.xlsx')
    print(wb.sheetnames)
    for sh in wb:
        print(sh.title)
def one_value():
    from openpyxl import load_workbook
    wb = load_workbook('test1.xlsx')
    sh = wb['Sheet1']
    print(sh['A1'].value)
    print(sh['A2'].value)
    print(sh['A3'].value)

def all_vaule():
    from openpyxl import load_workbook
    wb = load_workbook('test1.xlsx')
    sh = wb['Sheet1']
    for i in sh.rows:
        for j in i:
            print(j.value,end=' ')
        print()

    for i in sh.columns:
        for j in i:
            print(j.value,end=' ')
        print()
all_vaule()