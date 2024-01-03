# -*- coding: utf-8 -*-
# @Time    : 2022/10/23 22:36
# @Author  : cxt
# @Email   : 3026477511@qq.com
# @File    : douban.py
# @Software: PyCharm
# 大概就是我需要爬取豆瓣电影top250的电影名，导演，评分和经典语录，把爬出来的内容保存为xlsx文件.
import openpyxl
import requests
from bs4 import BeautifulSoup

num = 0

headers = {
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
    "Accept-Encoding": "gzip, deflate",
    "Accept-Language": "zh-CN,zh;q=0.9",
    "Cache-Control": "max-age=0",
    "Connection": "keep-alive",
    "Cookie": 'll="118200"; bid=if7o-u5rpLw; _vwo_uuid_v2=D42BF118CA0CFBF33DE8D94435B90D555|8925bd17a31f8b7058e92337ffcb5018; __gads=ID=4dc7fd39de72a642-22ad1e0315d700a5:T=1666015161:RT=1666015161:S=ALNI_MYaFQPlEV9oanRAwTzfwojslL387A; dbcl2="263719242:cdu+tcWOYSA"; __utmz=30149280.1666020617.4.3.utmcsr=accounts.douban.com|utmccn=(referral)|utmcmd=referral|utmcct=/; __utmz=223695111.1666020617.3.2.utmcsr=accounts.douban.com|utmccn=(referral)|utmcmd=referral|utmcct=/; push_noty_num=0; push_doumail_num=0; ct=y; __yadk_uid=YjGciSzXayZFWFJ5s9NkPcW5sv8CoJWw; ck=c-17; _pk_ref.100001.4cf6=%5B%22%22%2C%22%22%2C1666700334%2C%22https%3A%2F%2Faccounts.douban.com%2F%22%5D; __utma=30149280.1094459905.1662650904.1666623522.1666700335.9; __utmc=30149280; __utma=223695111.1141913362.1662650905.1666623522.1666700335.8; __utmc=223695111; __gpi=UID=00000b6478645c88:T=1666015161:RT=1666700335:S=ALNI_MarpFDTt2lvlUo2-EkBJy3V-uPMsA; _pk_id.100001.4cf6=6be088305e22e180.1662650904.8.1666700967.1666625362.',
    "Host": "movie.douban.com",
    "sec-ch-ua": "\"Chromium\";v=\"106\", \"Google Chrome\";v=\"106\", \"Not;A=Brand\";v=\"99\"",
    "sec-ch-ua-mobile": "?0",
    "sec-ch-ua-platform": "\"Windows\"",
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "none",
    "Sec-Fetch-User": "?1",
    "Upgrade-Insecure-Requests": "1",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/106.0.0.0 Safari/537.36"
}
# 定义四个列表,提取top250的电影名，导演，评分和经典语录
movie_names = []
directors = []
scores = []
comments = []
for i in range(10):
    url = f'https://movie.douban.com/top250?start={num}&filter='
    response = requests.get(url, headers=headers)
    # print(response.text)
    soup = BeautifulSoup(response.text, 'html.parser')
    title_list = soup.find_all('span', class_='title')
    directors_list = soup.find('p', attrs={"class": ""})
    scores_list = soup.find_all('span', class_="rating_num")
    for title in title_list:  # 获取标题
        # print('标题：|',title.text)
        # print('标题：|',title.string)
        if title.text.startswith('\xa0/\xa0'):
            pass
        else:
            movie_names.append(title.string)
    for movie in soup.select('.item'):
        director_list = movie.find('p', attrs={"class": ""})
        if director_list:
            my_director = director_list.get_text().strip().split(' ')[1]
            directors.append(my_director)
    for score in scores_list:
        scores.append(score.text)
    for movie in soup.select('.item'):
        if movie.select('.bd .quote') == []:
            comments.append('空')
        else:
            comment = movie.select('.bd .quote')[0].text  # 影评
            comments.append(comment.strip())
    num += 25
# print(movie_names)
# print(directors)
# print(scores)
# print(comments)
try:
    woorbook = openpyxl.load_workbook('信息.xlsx')  # 存在打开
except Exception as e:
    woorbook = openpyxl.Workbook()  # 不存在就创建
sheet_name = woorbook.active
# 写入表头信息
my_title = ['电影名称', '导演', '评分', '评论']
sheet_name.append(my_title)
# 把电影名称写入第一列
for i in range(4):
    if i==0:
        for j, value in enumerate(movie_names):
            sheet_name.cell(row=j + 2, column=i + 1, value=value)
    if i==1:
        for m, value in enumerate(directors):
            sheet_name.cell(row=m + 2, column=i + 1, value=value)
    if i==2:
        for n, value in enumerate(scores):
            sheet_name.cell(row=n + 2, column=i + 1, value=value)
    if i==3:
        for o, value in enumerate(comments):
            sheet_name.cell(row=o + 2, column=i + 1, value=value)
woorbook.save('信息.xlsx')
woorbook.close()