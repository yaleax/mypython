
def new():
    from openpyxl import workbook
    wb = workbook.Workbook()
    sh1 = wb.active
    sh2 = wb.create_sheet('动漫')
    wb.save('test2.xlsx')

def write():
    from openpyxl import load_workbook
    wb = load_workbook('test2.xlsx')
    sh = wb['动漫']
    sh['A1'] = '火影忍者'
    sh['A2'] = '海贼王'
    sh['A3'] = '死神'
    wb.save('test2.xlsx')

def write2():
    from openpyxl import workbook
    wb = workbook.Workbook()
    sh = wb.active
    rows = [['犬夜叉','银魂'],['火影忍者','海贼王'],['死神','妖精的尾巴']]
    for i in rows:
        sh.append(i)
        print(i)
    wb.save('test3.xlsx')
write2()

