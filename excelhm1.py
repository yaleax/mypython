
def new():
    from openpyxl import workbook
    wb = workbook.Workbook()
    sh = wb.active
    rows = [
        ['不良人', '李星云', 9.8, '国产'],
        ['堀与宫村', '堀京子', 9.6, '日本'],
        ['咒术回战', '虎杖', 9.7, '日本'],
        ['西游记之再世妖王', '孙悟空', 8.3, '国产'],
        ['冰雪大作战', '苏菲', 9.1, '加拿大'],
    ]
    for i in rows:
        sh.append(i)
        print(i)
    wb.save('test4.xlsx')
#new()
def check():
    from openpyxl import load_workbook
    wb = load_workbook('test4.xlsx')
    sh = wb.active
    rows = sh.rows
    f = 0 # 记录总分
    g = 0 # 记录总个数
    for row in rows:
        if float(row[2].value) >= 9:
            f += 1
        if row[3].value == '日本':
            g += 1
    tb = wb.create_sheet('统计')
    tb.append(["评分超9的", "国产"])
    tb.append([f, g])
    wb.save('test4.xlsx')

check()
